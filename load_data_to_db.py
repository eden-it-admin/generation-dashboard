# load_data_to_db.py
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import sqlite3
import os
import re

# --- Configuration ---
DATABASE_FILE = "plant_database.db"
DATA_FOLDER = "data_files/"
SHEET_NAME = "All Site Data"
TABLE_NAME_EXCEL = "Data" # Name of the table *within* the Excel sheet
# --- End Configuration ---

# Function to sanitize table names for SQL (replace spaces, special chars)
def sanitize_table_name(name):
    # Remove special characters except underscore
    s = re.sub(r'[^\w\s]', '', name)
    # Replace spaces with underscores
    s = re.sub(r'\s+', '_', s)
    # Ensure it starts with a letter or underscore (important for SQL)
    if not re.match(r'^[a-zA-Z_]', s):
        s = '_' + s
    return s

# Function to read specific table from Excel (copied from your original script)
def read_excel_tables(file_path, sheet_name, table_name):
    try:
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in {file_path}")
            return None
        current_sheet = wb[sheet_name]

        target_table = None
        for table in current_sheet.tables.values():
             if table.name == table_name:
                 target_table = table
                 break # Found the table

        if target_table:
            data_range = target_table.ref
            # Extract header range (first row of the table)
            header_start_cell = data_range.split(':')[0]
            header_row = int(re.search(r'\d+', header_start_cell).group()) # Get row number
            # Read using header row number (0-indexed, so subtract 1)
            # usecols is tricky with openpyxl table refs, read full then select is safer
            df = pd.read_excel(
                file_path, sheet_name=sheet_name, header=header_row-1 # 0-indexed header
            )
            # Now filter df to only include columns within the table's range
            start_col_idx = current_sheet[data_range.split(':')[0]].column - 1 # 0-indexed
            end_col_idx = current_sheet[data_range.split(':')[1]].column - 1 # 0-indexed
            # Filter rows based on table range (skip header row read by pandas)
            start_row_idx = int(re.search(r'\d+', data_range.split(':')[0]).group()) # 1-based table start
            end_row_idx = int(re.search(r'\d+', data_range.split(':')[1]).group())   # 1-based table end
            # Select relevant columns and rows (adjusting for pandas 0-based indexing and header row)
            df = df.iloc[start_row_idx-header_row : end_row_idx-header_row+1, start_col_idx : end_col_idx+1]
            # Reset index if needed
            df.reset_index(drop=True, inplace=True)
            return df
        else:
             print(f"Warning: Table '{table_name}' not found in sheet '{sheet_name}' of {file_path}")
             return None
    except Exception as e:
        print(f"Error reading Excel table from {file_path}: {e}")
        return None

# --- Main Loading Logic ---
print("Starting data loading process...")

# Find all Excel files
excel_files = [f for f in os.listdir(DATA_FOLDER) if f.endswith(".xlsx")]
if not excel_files:
    print(f"No Excel files found in '{DATA_FOLDER}'. Exiting.")
    exit()

print(f"Found {len(excel_files)} Excel files: {excel_files}")

# Connect to SQLite database (will create if not exists)
conn = sqlite3.connect(DATABASE_FILE)
cursor = conn.cursor()
print(f"Connected to database: {DATABASE_FILE}")

# Process each Excel file
for file in excel_files:
    file_path = os.path.join(DATA_FOLDER, file)
    print(f"\nProcessing file: {file_path}...")

    df_excel = read_excel_tables(file_path, SHEET_NAME, TABLE_NAME_EXCEL)

    if df_excel is None or df_excel.empty:
        print(f"Skipping file {file} due to read error or no data.")
        continue

    # **Data Cleaning and Type Conversion (Crucial before DB insertion)**
    try:
        # Handle Percentage Columns (convert to numeric 0-1 scale)
        for col in df_excel.columns:
            if isinstance(col, str) and ("%" in col or "percentage" in col.lower()):
                 # Check if it's the Soil Loss column - keep as string if so
                 if "Soil Loss(%)" in col or "Soil loss(%)" in col:
                     df_excel[col] = df_excel[col].astype(str)
                     print(f"  Keeping column '{col}' as string.")
                     continue # Skip numeric conversion for this specific column

                 # For other percentage columns:
                 if df_excel[col].dtype == 'object': # Only process if it's potentially string-like
                    df_excel[col] = df_excel[col].astype(str).str.replace('%', '', regex=False).str.strip()
                    # Replace non-numeric placeholders if any before conversion
                    df_excel[col] = df_excel[col].replace(['-', 'NA', 'N/A', ''], None) # Replace common non-numeric markers
                 # Convert to numeric, coercing errors
                 df_excel[col] = pd.to_numeric(df_excel[col], errors='coerce')
                 # Divide by 100 only if values seem like percentages (e.g. > 1)
                 # This check might need adjustment based on your actual data
                 if df_excel[col].notna().any() and (df_excel[col].dropna() > 1).any():
                      df_excel[col] = df_excel[col] / 100.0
                 print(f"  Converted percentage column '{col}' to numeric (0-1 scale).")


        # Handle Numeric Columns (like Generation MWHr)
        for col in df_excel.columns:
             if isinstance(col, str) and ("(MWHr)" in col or "(MW)" in col or "Capacity" in col):
                  # Check if it's already numeric, if not try conversion
                  if not pd.api.types.is_numeric_dtype(df_excel[col]):
                      df_excel[col] = pd.to_numeric(df_excel[col], errors='coerce')
                      print(f"  Converted column '{col}' to numeric.")

        # Handle Date Column ('Months')
        if 'Months' in df_excel.columns:
            df_excel['Months'] = pd.to_datetime(df_excel['Months'], errors='coerce')
            # Store dates as ISO format strings (YYYY-MM-DD), good for SQLite TEXT columns
            df_excel['Months'] = df_excel['Months'].dt.strftime('%Y-%m-%d')
            print("  Converted 'Months' column to YYYY-MM-DD string format.")

        # Fill NaNs created during conversion if necessary (optional, depends on needs)
        # df_excel.fillna(0, inplace=True) # Example: fill numeric NaNs with 0

    except Exception as e:
        print(f"  Error during data cleaning/conversion for {file}: {e}")
        continue # Skip this file if cleaning fails

    # Get unique Plant names from this file's data
    if 'Plant' not in df_excel.columns:
        print(f"  Error: 'Plant' column not found in {file}. Cannot determine target table. Skipping.")
        continue

    unique_plants = df_excel['Plant'].unique()
    print(f"  Found plants in this file: {unique_plants}")

    # Insert data into respective plant tables
    for plant_name in unique_plants:
        if pd.isna(plant_name):
            print("  Skipping rows with missing Plant name.")
            continue

        plant_table_name = sanitize_table_name(str(plant_name))
        df_plant_data = df_excel[df_excel['Plant'] == plant_name].copy()

        # Drop rows where 'Months' became NaT/None after conversion, if any
        df_plant_data.dropna(subset=['Months'], inplace=True)

        if df_plant_data.empty:
            print(f"  No valid data found for plant '{plant_name}' after cleaning.")
            continue

        print(f"    Writing {len(df_plant_data)} rows for plant '{plant_name}' to table '{plant_table_name}'...")
        try:
            # Use 'replace' to overwrite table if it exists, or 'append' to add
            df_plant_data.to_sql(plant_table_name, conn, if_exists='replace', index=False)
            print(f"    Successfully wrote data to table '{plant_table_name}'.")
        except Exception as e:
            print(f"    Error writing to table '{plant_table_name}': {e}")
            # Optional: Log more details about the specific data causing issues
            # print(df_plant_data.info())
            # print(df_plant_data.head())


# Close the database connection
conn.commit() # Commit changes
conn.close()
print("\nDatabase loading process finished.")